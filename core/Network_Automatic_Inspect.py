#!/usr/bin/env python3
# -*- coding:utf-8 -*-

"""
作者：by singvis
微信公众号: 点滴技术
B站：点滴技术
update： 2020.08.03
1.优化下打印输出
2.windows环境格式化文件命令(replace特殊符号)

update:2022.08.24
1.增加华为conn_timeout参数

update:2022.09.18
1.更新disconnect()关闭会话

欢迎大家关注，点赞 收藏 分享，3连击.
"""

import os
import re
import threading
import platform

from pathlib import Path
from backend import settings
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from multiprocessing.pool import ThreadPool
from prettytable import PrettyTable
from netmiko import ConnectHandler
from netmiko.ssh_exception import (
    NetMikoTimeoutException,
    AuthenticationException,
    SSHException
)
from netmiko.fortinet import FortinetSSH
from netmiko.juniper import JuniperSSH

BASE_DIR = Path(__file__).resolve().parent

RE_HOSTNAME = {
    'huawei': re.compile(r"(?<=(\<|\[)).*?(?=(\>|\]))", re.IGNORECASE),  # <hostname> or [hostname]
    'hp_comware': re.compile(r"(?<=(\<|\[)).*(?=(\>|\]))", re.IGNORECASE),  # <hostname> or [hostname]
    'cisco': re.compile(r".*?(?=(>|#))", re.IGNORECASE),  # hostname> or hostname#
    'aruba': re.compile(r".*?(?=(>|#))", re.IGNORECASE),  # hostname#
    'fortinet': re.compile(r".*?(?=#)", re.IGNORECASE),  # hostname#
    'a10': re.compile(r".*?(?=(>|#))", re.IGNORECASE),  # hostname-Active> or hostname-Active#
    'paloalto': re.compile(r"(?<=(@)).*?(?=(\(|\>))", re.IGNORECASE),  # admin@hostname(active)>,
    'juniper': re.compile(r".*?(?=(\-\>))", re.IGNORECASE),  # hostname->
}


class BaseConnection(object):
    """
    基类
    """
    pass


class MyFortinetSSH(FortinetSSH):
    """
    重写了FortinetSSH类
    """

    # 因为通过ftp备份配置，不需要关闭分屏
    # 多vdom场景下，需要进入global模式，需要有一些权限权限，可以备份多个vdom的配置
    def disable_paging(self, delay_factor=1, **kwargs):
        check_command = "get system status | grep Virtual"
        output = self.send_command_timing(check_command)
        self.allow_disable_global = True
        self.vdoms = False
        self._output_mode = "more"

        if re.search(r"Virtual domain configuration: (multiple|enable)", output):
            self.vdoms = True
            vdom_additional_command = "config global"
            output = self.send_command_timing(vdom_additional_command, delay_factor=2)
            if "Command fail" in output:
                self.allow_disable_global = False
                self.remote_conn.close()
                self.establish_connection(width=100, height=1000)
        return output


class MyJuniperSSH(JuniperSSH):
    """重写了JuniperSSH类"""

    # netscreen 不支持""set cli screen-width 511""命令，调整下命令
    def session_preparation(self):
        """Prepare the session after the connection has been established."""
        self.enter_cli_mode()

        self.disable_paging(
            command="set console page 0", pattern=r"->"
        )
        self.set_base_prompt()


class NetworkHandler(object):
    def __init__(self):
        """初始参数"""
        self.BaseDir = BASE_DIR.parent  # 主入口文件夹路径
        self.device_file = self.BaseDir / "templates" / "巡检模板.xlsx"  # 巡检模板文件，文件名固定的
        self.pool = ThreadPool(50)  # 并发数，可按需修改
        self.queueLock = threading.Lock()  # 线程锁
        self.logtime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # 时间
        self.log = self.log_dir()
        self.FtpServer = settings.ftp_server
        self.FtpUser = settings.ftp_user
        self.FtpPassword = settings.ftp_password

        self.success = []
        self.fail = []

    def log_dir(self):
        """创建目录"""
        try:
            LOG = self.BaseDir / 'LOG'
            # 判断当前目录是否有LOG文件夹，不存在则创建
            if not os.path.exists(LOG):
                os.makedirs(LOG)
            return LOG

        except FileNotFoundError:
            raise "{}".format("LOG文件夹创建失败.")

    def printPretty(self, msg):
        """打印消息"""
        # 在并发的场景中，避免在一行打印出多个结果，不方便查看
        self.queueLock.acquire()  # 加锁
        print(msg)
        self.queueLock.release()  # 释放锁

    def printSum(self, msg):
        """打印结果汇总信息"""
        total_devices, success, fail = len(self.success + self.fail), len(self.success), len(self.fail)
        total_time = "{:0.2f}s".format(msg.total_seconds())
        tb = PrettyTable(['设备总数', '成功', '失败', '总耗时'])
        tb.add_row([total_devices, success, fail, total_time])
        print(tb)

    def write_to_file(self, *args, **kwargs):
        """
        将结果写入文件，action的序号保持和choice_function对应，方便理解
        action = 1, 表示连接设备测试的动作
        action = 2, 表示采集设备配置的动作
        action = 3, 表示保存设备配置的动作
        action = 4, 表示下发设备配置的动作
        """
        try:
            if kwargs['action'] == 1:
                # 连接设备测试的动作,将结果写入文件
                with open(self.log / f'connect_t_{self.logtime}.log', 'a') as f:
                    f.write(kwargs['result'])
                    f.write('\n')
            elif kwargs['action'] == 2:
                # 采集设备配置的动作
                if kwargs['code'] == 1:
                    # 将异常的结果写入文本
                    with open(os.path.join(self.log, f'error_{self.logtime}.log'), 'a') as f:
                        f.write(kwargs['result'])
                        f.write('\n')
                else:
                    # 将正常的结果写入文本
                    # LOG/2022-01-01_00:00:01/192.168.1.1_Router01/show run.conf
                    with open(kwargs['path'], 'a') as f:
                        f.write(kwargs['result'])
            elif kwargs['action'] == 3:
                # 保存设备配置的动作
                if kwargs['code'] == 1:
                    # 将异常的结果写入文本
                    with open(os.path.join(self.log, f'error_{self.logtime}.log'), 'a') as f:
                        f.write(kwargs['result'])
                        f.write('\n')
                else:
                    # 将正常的结果写入文本
                    # LOG/2022-01-01_00:00:01/192.168.1.1_Router01/show run.conf
                    with open(kwargs['path'], 'a') as f:
                        f.write(kwargs['result'])
            else:
                # 将其他异常的写入文件
                with open(os.path.join(self.log, f'error_{self.logtime}.log'), 'a') as f:
                    f.write(kwargs['result'])
                    f.write('\n')
        except Exception as e:
            self.printPretty(f"action:{kwargs['action']}错误: {e}")

    def load_excel(self):
        """加载excel文件"""
        try:
            wb = load_workbook(self.device_file)
            return wb

        except FileNotFoundError:
            raise ("{} 文件不存在.".format(self.device_file))

    def get_devices_info(self, action=0):
        """获取设备的基本信息"""
        try:
            wb = self.load_excel()
            ws1 = wb[wb.sheetnames[0]]

            n = 0
            # 参数min_row、max_col限制取值范围
            # row(行)从id=1开始...
            # col(列)从id=1开始...
            for row in ws1.iter_rows(min_row=2, max_col=10):
                n += 1
                if str(row[2].value).strip() == '#':
                    # 跳过注释行
                    continue
                info_dict = {
                    'ip': row[3].value,
                    'protocol': row[4].value,
                    'port': row[5].value,
                    'username': row[6].value,
                    'password': row[7].value,
                    'secret': row[8].value,
                    'device_type': row[9].value,
                    'cmd_list': self.get_cmd_info(wb[row[9].value]) if row[9].value else '',
                    'session_log': row[3].value + '.log',
                    'session_log_record_writes': True
                }

                yield info_dict

        except Exception as e:
            output = f"Excel_第{n}行_Error: {e}"
            self.printPretty(output)
            self.write_to_file(**{'action': action, 'result': str(output)})
        else:
            # 记得最后要关闭workbook
            wb.close()

    def get_cmd_info(self, cmd_sheet):
        """获取命令的信息"""
        cmd_list = []
        try:
            for row in cmd_sheet.iter_rows(min_row=2, max_col=2):
                if str(row[0].value).strip() != "#" and row[1].value:
                    # 跳过注释行，去掉命令左右空白处
                    cmd_list.append(row[1].value.strip())

            return cmd_list

        except Exception as e:
            self.printPretty("get_cmd_info Error: {}".format(e))

    def format_hostname(self, hostname, device_type):
        """格式化主机名称"""
        try:
            for vendor, regex in RE_HOSTNAME.items():
                if vendor in device_type:
                    match = re.search(regex, hostname)
                    if match:
                        new_hostname = match.group(0)
                    else:
                        new_hostname = hostname.split()[0].strip("<>#$()[] ")

        except Exception as e:
            self.printPretty(e)
            raise e

        return new_hostname

    def format_cmd(self, cmd):
        # 格式化命令行
        # 避免windown环境文件命令不允许特殊符号,按需修改
        if platform.system().lower() == 'windows':
            cmd = cmd.replace('|', '_')
        else:
            cmd = cmd
        return cmd

    def connectHandler(self, host, action=None):
        """定义一个netmiko对象"""
        try:
            connect = ''

            # 判断使用ssh协议
            if host['protocol'].lower().strip() == 'ssh':
                host['port'] = host['port'] if (host['port'] not in [22, None]) else 22
                host.pop('protocol'), host.pop('cmd_list')

                if 'huawei' in host['device_type']:
                    connect = ConnectHandler(**host, conn_timeout=15)
                elif 'fortinet' in host['device_type']:
                    # 调用重写的MyFortinetSSH类
                    connect = MyFortinetSSH(**host)
                elif 'juniper' in host['device_type']:
                    # 优化netscreen设备(优化分屏命令)
                    connect = MyJuniperSSH(**host)
                else:
                    connect = ConnectHandler(**host)
            # 判断使用telnet协议
            elif host['protocol'].lower().strip() == 'telnet':
                host['port'] = host['port'] if (host['port'] not in [23, None]) else 23
                host.pop('protocol'), host.pop('cmd_list')
                # netmiko里面支持telnet协议，示例：cisco_ios_telnet
                host['device_type'] = host['device_type'] + '_telnet'

                # fast_cli=False，为了修复telnet login authentication 报错.
                connect = ConnectHandler(**host, fast_cli=False)
            else:
                # 不支持的协议
                raise ValueError("{}协议格式填写错误!".format(host['protocol']))

            return connect

        # 异常捕获
        except NetMikoTimeoutException:
            e = "Failed.....{:<15} 连通性问题!".format(host['ip'])
            self.printPretty(e)
            self.fail.append(host['ip'])
            self.write_to_file(**{'action': action, 'code': 1, 'result': str(e)})
        except AuthenticationException:
            e = "Failed.....{:<15} 用户名或密码错误!".format(host['ip'])
            self.printPretty(e)
            self.fail.append(host['ip'])
            self.write_to_file(**{'action': action, 'code': 1, 'result': str(e)})
        except SSHException:
            e = "Failed.....{:<15} SSH版本不兼容!".format(host['ip'])
            self.printPretty(e)
            self.fail.append(host['ip'])
            self.write_to_file(**{'action': action, 'code': 1, 'result': str(e)})
        except Exception as e:
            e = "Failed.....{:<15} connectHandler Error: {}".format(host['ip'], e)
            self.printPretty(e)
            self.fail.append(host['ip'])
            self.write_to_file(**{'action': action, 'code': 1, 'result': str(e)})

    def _threadpool(self, *args, **kwargs):
        # hosts是一个生成器，需要for循环进行遍历
        hosts = self.get_devices_info()
        for host in hosts:
            # 多进程并发
            self.pool.apply_async(
                func=kwargs["func"],
                args=(host,)
            )
        self.pool.close()
        self.pool.join()

    def run_t(self, host, action=1):
        """主要获取设备名称提示符"""
        conn = self.connectHandler(host, action=action)
        if conn:
            try:
                output = "获取设备的提示符: {}".format(conn.find_prompt())

                self.printPretty(output)
                self.success.append(host['ip'])  # 追加到成功的列表
                self.write_to_file(**{'action': action, 'result': output})  # 将结果接入文件

                # 最后一定要关闭会话
                conn.disconnect()

            except Exception as e:
                output = f"run_t Failed...{host['ip']} : {e}"
                self.printPretty(output)
                self.fail.append(host['ip'])
                self.write_to_file(**{'action': action, 'result': output})
            finally:
                # 退出netmiko session
                conn.disconnect()

    def run_get_config(self, host, cmds, action=2):
        """执行命令和保存信息"""
        self.printPretty('设备...{:.<15}...开始执行'.format(host['ip']))

        # 特权功能标识位
        enable = True if host['secret'] else False

        conn = self.connectHandler(host, action=action)

        if conn:
            # 这里要注意下windown文件命名不能有特殊符号，否则会创建失败
            try:
                # 获取设备名称并格式化
                hostname = self.format_hostname(conn.find_prompt(), host['device_type'])
                dirname = host['ip'] + '_' + hostname  # 192.168.1.1_Router-01
                dirpath = os.path.join(self.log, self.logtime, dirname)  # LOG/2022-01-01_00:00:01/192.168.1.1_Router01
                # 递归创建目录
                os.makedirs(dirpath)
            except Exception as e:
                raise Exception("文件夹创建失败!{}".format(e))

            try:
                if cmds:
                    for cmd in cmds:
                        if enable:
                            # 进入特权模式
                            if 'cisco' or 'ruijie' in host['device_type']:
                                # 默认源码只有cisco支持enabel命令，国产的super需要改写
                                conn.enable()
                                output = conn.send_command(cmd)
                                data = {'action': action, 'code': 0, 'result': output,
                                        'path': os.path.join(dirpath, self.format_cmd(cmd) + '.conf')}
                                self.write_to_file(**data)
                            else:
                                # 留空，拓展其他厂商
                                pass
                        else:
                            output = conn.send_command(cmd)
                            data = {'action': action, 'code': 0, 'result': output,
                                    'path': os.path.join(dirpath, self.format_cmd(cmd) + '.conf')}
                            self.write_to_file(**data)
                else:
                    # 适用于ftp/sftp/scp备份
                    if host['device_type'] == 'fortinet':
                        # 飞塔防火墙FTP备份
                        cmd = "execute backup config ftp {}_{}.conf {} {} {}".format(hostname, self.logtime,
                                                                                     self.FtpServer,
                                                                                     self.FtpUser, self.FtpPassword)
                        conn.send_command(cmd, expect_string="to ftp server OK")
                    elif host['device_type'] == 'cisco_wlc':
                        # 按需补充
                        pass
                    else:
                        pass

                self.success.append(host['ip'])

            except Exception as e:
                output = f"run Failed...{host['ip']} : {e}"
                self.printPretty(output)
                self.fail.append(host['ip'])
                self.write_to_file(**{'action': action, 'code': 1, 'result': str(output)})

            finally:
                # 退出netmiko session
                conn.disconnect()

    def run_save_config(self, host, action=3):
        """
        执行保存网络设备配置
        """
        self.printPretty('设备...{:.<15}...开始执行'.format(host['ip']))
        # 特权功能标识位
        enable = True if host['secret'] else False

        conn = self.connectHandler(host, action=action)

        if conn:
            try:
                # 获取设备名称并格式化
                hostname = self.format_hostname(conn.find_prompt(), host['device_type'])
                dirname = host['ip'] + '_' + hostname  # 192.168.1.1_Router-01
                dirpath = os.path.join(self.log, self.logtime, dirname)  # LOG/2022-01-01_00:00:01/192.168.1.1_Router01
                # 递归创建目录
                os.makedirs(dirpath)
            except Exception as e:
                raise Exception("文件夹创建失败!{}".format(e))

            try:
                # 后台会自动进入特权模式，并执行设备保存命令
                output = conn.save_config()
                # self.printPretty(f"{host['ip']}_设备的操作记录:\n{output}")
                self.success.append(host['ip'])
                data = {
                    'action': action,
                    'code': 0,
                    'result': output,
                    'path': os.path.join(dirpath, hostname + '.log')
                }
                self.write_to_file(**data)

            except Exception as e:
                output = f"execute save config Failed...{host['ip']} : {e}"
                self.printPretty(f"报错: {output}")
                self.fail.append(host['ip'])
                self.write_to_file(**{'action': action, 'code': 1, 'result': output})

    def run_send_config(self, host=None, action=4):
        pass

    def main_connect_t(self):
        """
        主程序，测试登录设备
        """
        # 开始时间
        start_time = datetime.now()
        # 调用进程池
        self._threadpool(func=self.run_t)

        # 结束时间
        end_time = datetime.now()
        # 打印结果
        self.printSum(end_time - start_time)

    def main_get_config(self):
        """
        主程序，采集设备配置信息
        """
        start_time = datetime.now()

        # hosts 是一个生成器，需要for循环进行遍历
        hosts = self.get_devices_info()
        for host in hosts:
            self.pool.apply_async(self.run_get_config, args=(host, host['cmd_list']))
        self.pool.close()
        self.pool.join()

        # 结束时间
        end_time = datetime.now()
        # 打印结果
        self.printSum(end_time - start_time)

    def main_save_config(self):
        """
        主程序，保存设备配置
        """
        # 开始时间
        start_time = datetime.now()

        # hosts 是一个生成器，需要for循环进行遍历
        hosts = self.get_devices_info()
        for host in hosts:
            self.pool.apply_async(self.run_save_config, args=(host,))
        self.pool.close()
        self.pool.join()

        # 结束时间
        end_time = datetime.now()
        # 打印结果
        self.printSum(end_time - start_time)

    def main_send_config(self):
        """
        主程序，保存设备配置
        """
        # 开始时间
        start_time = datetime.now()

        # hosts 是一个生成器，需要for循环进行遍历
        hosts = self.get_devices_info()
        for host in hosts:
            self.pool.apply_async(self.run_send_config, args=(host,))
        self.pool.close()
        self.pool.join()

        # 结束时间
        end_time = datetime.now()
        # 打印结果
        self.printSum(end_time - start_time)


if __name__ == '__main__':
    # pass
    n = NetworkHandler()
    n.main_connect_t()

