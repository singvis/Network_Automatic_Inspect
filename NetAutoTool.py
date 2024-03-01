#!/usr/bin/env python3
# -*- coding:utf-8 -*-
"""
Author：Singvis
微信公众号: 点滴技术
B站：点滴技术
"""

import os
import json
import logging
import platform
import re
import nmap
import subprocess
import gevent.pool

from gevent import monkey

# monkey补丁
monkey.patch_all()

try:
    import netmiko

    if netmiko.__version__ != "3.4.0":
        print("建议将Netmiko版本安装为3.4.0")
except ImportError:
    print("未安装Netmiko模块")
except Exception as err:
    print(f"Netmiko报错: {err}")
from config import global_config
from typing import Union
from functools import wraps
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from prettytable import PrettyTable
from extras.ssh_dispatcher import (FortinetSSH, JuniperSSH, FiberHomeSSH, HillStoneSSH, MaipuSSH)
from extras.snmp_autodetect import MySNMPDetect
from extras.ssh_autodetect import MySSHDetect
from jinja2 import FileSystemLoader, Environment
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoAuthenticationException, NetmikoTimeoutException
from netmiko.ssh_autodetect import SSHDetect
from gevent.lock import BoundedSemaphore
from utils.RichTool import RichTool
from utils.ping import Ping

RE_HOSTNAME = {
    'huawei': re.compile(r"(?<=(\<|\[)).*?(?=(\>|\]))", re.IGNORECASE),  # <hostname> or [hostname]
    'hp_comware': re.compile(r"(?<=(\<|\[)).*(?=(\>|\]))", re.IGNORECASE),  # <hostname> or [hostname]
    'cisco': re.compile(r".*?(?=(>|#))", re.IGNORECASE),  # hostname> or hostname#
    'aruba': re.compile(r".*?(?=(>|#))", re.IGNORECASE),  # hostname#
    'fortinet': re.compile(r".*?(?=#)", re.IGNORECASE),  # hostname#
    'a10': re.compile(r".*?(?=(>|#))", re.IGNORECASE),  # hostname-Active> or hostname-Active#
    'paloalto': re.compile(r"(?<=(@)).*?(?=(\(|\>))", re.IGNORECASE),  # admin@hostname(active)>,
    'juniper': re.compile(r".*?(?=(\-\>))", re.IGNORECASE),  # hostname->
    'linux': re.compile(r"(?<=(\[)).*?(?=(~|]))", re.IGNORECASE),
}

RE_VENDOR = {
    "huawei": re.compile('huawei', re.IGNORECASE),
    "h3c": re.compile('hp_comware', re.IGNORECASE),
    'cisco': re.compile('cisco', re.IGNORECASE),
    'aruba': re.compile('aruba', re.IGNORECASE),
    'a10': re.compile('a10', re.IGNORECASE),
    'fortinet': re.compile('fortinet', re.IGNORECASE),
    'paloalto': re.compile('paloalto', re.IGNORECASE),
    'juniper': re.compile('juniper', re.IGNORECASE),
}


def async_task(wrapped):
    """
    装饰器
    """
    @wraps(wrapped)
    def wrapper(self, *args, **kwargs):
        start_time = datetime.now()
        # print(f"args:{args}, kwargs:{kwargs}")
        greenlets = [self.async_pool.spawn(wrapped, self, arg) for arg in args[0]]
        gevent.joinall(greenlets)

        end_time = datetime.now()
        print("总共耗费时长 {:0.2f} 秒".format((end_time - start_time).total_seconds()))
        self.printSum((end_time - start_time).total_seconds())

    return wrapper


class NetAutoTool(object):
    def __init__(self):
        """初始参数"""
        # 基础信息
        self.device_file = "巡检模板.xlsx"  # 模板文件
        self.device_port = list(global_config.get('nmap', 'scan_device_port').split(','))  # 扫描设备端口，缺省22,23
        self.username = global_config.get('account', 'username')  # 用户名
        self.password = global_config.get('account', 'password')  # 用户密码
        # self.config_username = global_config.get('account', 'config_username')  # 管理用户
        # self.config_password = global_config.get('account', 'config_password')  # 管理用户密码
        self.secret = global_config.get('account', 'secret')  # 特权密码

        # 缓存数据
        self.cache_data_file = os.path.join('cache', 'cache_data.json')

        # log时间与目录
        self.logtime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")  # 日期时间
        self.log = "LOG"
        self.show_dir, self.config_dir, self.ping_dir, self.scan_dir = 'show', 'config', 'ping', "scan"
        self.fail_file = f"fail_{self.logtime}.log"

        # ftp服务器
        self.FtpServer = global_config.get('account', 'ftp_server')  # Ftp Server
        self.FtpUser = global_config.get('account', 'ftp_username')  # Ftp 用户名
        self.FtpPassword = global_config.get('account', 'ftp_password')  # Ftp 密码

        # 并发|锁
        self.async_pool = gevent.pool.Pool(int(global_config.get('config', 'concurrent_num')))
        self.geventLock = BoundedSemaphore(1)  # 携程锁

        # autodetect
        self.ssh_detect = global_config.get('netmiko', 'ssh_autodetect').lower()
        self.snmp_detect = global_config.get('netmiko', 'snmp_autodetect').lower()
        self.comunity = global_config.get('netmiko', 'snmp_community')

        # 成功与失败的设备列表
        self.success = []
        self.fail = []

        # 开关debug
        self.enableDebug()

        # rich
        self.rich = RichTool

    def enableDebug(self) -> None:
        if global_config.get('debug', 'debug').lower() == 'true':
            print("已开启debug模式...")
            logging.basicConfig(filename='debug.log', level=logging.DEBUG)
            logging.getLogger("netmiko")

    def printPretty(self, msg:str) -> str:
        """打印消息"""
        with self.geventLock:
            print(msg)

    def printSum(self, total_time) -> None:
        """打印结果汇总信息"""
        total_devices, success, fail = len(self.success + self.fail), len(self.success), len(self.fail)
        tb = PrettyTable(['设备总数', '成功', '失败', '总耗时'])
        tb.add_row([total_devices, success, fail, "{:0.2f}s".format(total_time)])
        print(tb)

    def time_now(self) -> str:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-4]

    def load_excel(self) -> load_workbook:
        """加载excel文件"""
        try:
            return load_workbook(self.device_file)
        except FileNotFoundError:
            raise FileNotFoundError("{} 文件不存在.".format(self.device_file))
        except Exception as e:
            raise f"load_excel_error: {e}."

    def get_devices_info(self, wb=None, row=None) -> dict:
        """
        通过nscan扫描出来的参数临时写入cache文件
        再回填到excel，可减少下scan动作
        """
        try:
            port = row[4].value if row[4].value else self.scan_port(ip=row[2].value, port=self.device_port)
            if not port:
                return {}  # 如果port扫描不通，则不往下执行
            info_dict = {
                'ip': row[2].value,
                'protocol': row[3].value,
                'port': port,
                'username': row[5].value if row[5].value else global_config.get('account', 'username').strip(),
                'password': row[6].value if row[6].value else global_config.get('account', 'password').strip(),
                'secret': row[7].value if row[7].value else global_config.get('account', 'secret').strip(),
                'template_file': row[9].value,
            }
            device_type = row[8].value if row[8].value else self.detect_deviceType(
                ip=row[2].value,
                username=info_dict.get("username"),
                password=info_dict.get('password'),
                secret=info_dict.get('secret'),
            )
            info_dict['cmd_list'] = self.get_cmd_info(wb, device_type)
            info_dict['device_type'] = device_type
            # print(info_dict)

            return info_dict

        except Exception as e:
            error = f"Excel__Error: {e}"
            raise ValueError(error)

    def get_devices_info_async(self) -> list[dict]:
        """
        如启用了nmap，通过异步方式减少scan等待时间
        return: 列表
        """
        wb = None
        row_num = None
        try:
            wb = self.load_excel()  # 工作薄对象
            first_sheet = wb[wb.sheetnames[0]]  # 获取第一个sheet
            results = []
            for row in first_sheet.iter_rows(min_row=2, max_col=10):  # 通过参数min_row/max_col限制区域范围
                row_num = row[0].row  # 行编号
                if row[1].value == '#':
                    continue  # 跳过注释行
                elif not self.validate_ip_address(row[2].value):
                    continue  # 跳过有误的IP地址格式
                else:
                    greenlet = gevent.spawn(self.get_devices_info, wb, row)
                    results.append(greenlet)

            # 等待所有线程完成
            gevent.joinall(results)

            # results= [<Greenlet at 0x20b16b96160: _run>, ...]
            # 组合判断ip、port和device_type必须为真条件(netmiko场景)
            device_info_list = [
                greenlet.value
                for greenlet in results
                if (greenlet.value and
                    greenlet.value.get("ip") and
                    greenlet.value.get("port") and
                    greenlet.value.get("device_type")) is not False
            ]
            # 写入缓存文件
            self.cache_data(device_info_list)

            return device_info_list

        except Exception as e:
            error = f"Excel_第{row_num}行_错误: {str(e)}"
            self.rich.errPrint(error)
            self.write_to_file(**{'code': 0, 'result': error, 'path': os.path.join(self.log, self.fail_file)})
            # raise ValueError(error)

        finally:
            wb.close()  # 记得最后要关闭workbook

    def validate_ip_address(self, ip: str) -> bool:
        """
        校验IPv4地址格式
        """
        try:
            ip_pattern = re.compile(r'^\d{1,3}(\.\d{1,3}){3}$')
            if ip_pattern.match(ip):
                return True
            else:
                return False
        except:
            output = "{:<20} IP地址格式不正确,请检查!".format(ip)
            self.rich.errPrint(output)
            self.write_to_file(**{'code': 0, 'result': str(output), 'path': os.path.join(self.log, self.fail_file)})
            return False

    def validate_ping_alive(self, ip: str) -> bool:
        """
        验证ip地址的ping连通性
        """
        try:
            ping_result = Ping.run_ping(ip)
            return ping_result[0]
        except:
            output = "{:<20} IP地址ping不通,请检查!".format(ip)
            self.rich.errPrint(output)
            self.write_to_file(**{'code': 0, 'result': str(output), 'path': os.path.join(self.log, self.fail_file)})
            return False

    def nscan(self) -> nmap.PortScanner():
        """nmap实例"""
        return nmap.PortScanner()

    def scan_port(self, *args, **kwargs) -> Union[str, bool]:
        """扫描设备端口"""
        try:
            ip = kwargs.get('ip')  # ip地址
            item = self.load_cache_data(ip)  # 加载缓存文件数据
            if item:
                return item.get("port")   # 如果缓存有，则直接返回数据
            else:
                self.rich.print("{} Start scanning port...{}".format(self.time_now(), ip))
                for port in kwargs.get('port'):
                    res = self.nscan().scan(ip, str(port))  # scan()的参数必须为字符串类型string
                    # 判断scan结果为真(不同nmap版本扫描结果会不太一样)
                    # 有的扫描端口不通，结果为空，有的扫描端口不通，结果是'filtered'
                    if res['scan']:
                        return port if res['scan'][ip]['tcp'][int(port)]['state'] == 'open' else None

                output = "{:<20} 端口扫描不可达,请检查!".format(ip)
                self.rich.errPrint(output)
                self.fail.append(ip)
                self.write_to_file(**{'code': 0, 'result': str(output), 'path': os.path.join(self.log, self.fail_file)})
                return False

        except Exception as e:
            self.rich.errPrint(str(e))
            return False

    def detect_deviceType(self, *args, **kwargs) -> Union[str, bool]:
        """检测设备类型"""
        try:
            ip = kwargs.get("ip")
            item = self.load_cache_data(ip)
            if item:
                return item.get("device_type")  # 如果缓存有，则直接返回数据
            else:
                self.rich.print("{} Start detecting device_type...{}".format(self.time_now(), ip))
                # snmp探测
                if self.snmp_detect == 'true' and self.comunity:
                    device_type = MySNMPDetect(
                        hostname=ip,
                        snmp_version="v2c",
                        community=self.comunity
                    ).autodetect()
                    # 如果snmp不通，返回的结果是None
                    if device_type:
                        return device_type
                    else:
                        # output = "{:<20} snmp不通,请检查!".format(ip)
                        # self.rich.errPrint(output)
                        # return False
                        raise ValueError("{:<20} SNMP: 设备类型(device_type)检测失败!".format(ip))
                # ssh探测
                if self.ssh_detect:
                    kwargs['device_type'] = "autodetect"
                    ssh = MySSHDetect(**kwargs)
                    device_type = ssh.autodetect()
                    # print(f"device_type: {device_type}")
                    if device_type:
                        return device_type
                    else:
                        raise ValueError("{:<20} SSH: 设备类型(device_type)检测失败!".format(ip))

        except Exception as e:
            self.rich.errPrint("设备类型检测错误: {}".format(str(e)))
            return False

    def write_to_file(self, *args, **kwargs) -> None:
        """
        将结果写入文件
        """
        try:
            with self.geventLock:
                with open(kwargs['path'], mode='a', encoding='utf-8') as f:
                    f.write(kwargs['result'] + "\n")
        except Exception as e:
            self.rich.errPrint(f"写入文件错误: {e}")

    def get_cmd_info(self, wb, sheet_name) -> list:
        """获取命令条目的信息"""
        cmd_list = []
        try:
            cmd_sheet = wb[sheet_name]
            for row in cmd_sheet.iter_rows(min_row=2, max_col=2):
                if str(row[0].value).strip() != "#" and row[1].value:
                    # 跳过注释行，去掉命令左右空白处
                    cmd_list.append(row[1].value.strip())
            return cmd_list

        except KeyError as e:
            raise KeyError(f"找不到sheet名称: {sheet_name}")

    def format_hostname(self, hostname, device_type) -> str:
        """格式化主机名称"""
        new_hostname = ''
        try:
            for vendor, regex in RE_HOSTNAME.items():
                if vendor in device_type:
                    match = re.search(regex, hostname)
                    if match:
                        new_hostname = match.group(0).strip()
                    else:
                        new_hostname = hostname.split()[0].strip("<>#$()[] ")

            return new_hostname

        except Exception as e:
            # self.printPretty(e)
            self.rich.errPrint(str(e))
            raise e

    def format_cmd(self, cmd) -> str:
        # 格式化命令行
        # windows文件命名不允许有特殊符号,按需修改
        if platform.system().lower() == 'windows':
            cmd = cmd.strip().replace('|', '_')
        else:
            cmd = cmd
        return cmd

    def normalize_netmiko(self, host) -> dict:
        """标准化netmiko参数"""
        new_host = {
            'ip': host["ip"],
            'port': host["port"],
            'username': host["username"],
            'password': host["password"],
            'secret': host["secret"],
            'device_type': host["device_type"],
        }
        return new_host

    def normalize_vendor(self, device_type) -> str:
        """标准化厂商名称"""
        try:
            for vendor, regex in RE_VENDOR.items():
                match = re.search(regex, device_type)
                if match:
                    new_vendor = vendor
                    return new_vendor
                else:
                    return device_type
        except Exception as e:
            raise ValueError(f"normalize_vendor错误: {e}.")

    def create_dir(self, dir_name) -> None:
        """
        创建目录
        """
        try:
            with self.geventLock:
                os.makedirs(os.path.join(self.log, dir_name), exist_ok=True)
        except Exception as e:
            output = f"创建目录失败: {e}"
            self.rich.errPrint(output)

    def render_jinjia2_tpl(self, vendor: str, template: str) -> str:
        """
        返回jinjia2模板内容
        """
        try:
            loader = FileSystemLoader(os.path.join("templates", 'config', vendor))
            env = Environment(loader=loader)
            tpl = env.get_template(f"{template}")
            return tpl.render()
        except Exception as e:
            raise ValueError(f"render_jinja2_err: {e}")

    def cache_data(self, data: list[dict]) -> None:
        """
        将自动探测的参数写入文件，作为缓存数据
        """
        # 仅缓存ip、port、device_type字段
        new_data = [{k: v for k, v in d.items() if k in ["ip", "port", "device_type"]} for d in data]

        with open(self.cache_data_file, mode='w', encoding='utf-8') as f:
            json.dump(new_data, f, indent=2)

    def load_cache_data(self, ip: str) -> Union[dict, None]:
        """
        读取设备的缓存数据
        """
        try:
            file_size = os.path.getsize(self.cache_data_file)  # 文件大小
            if not file_size:
                return None
            else:
                with open(self.cache_data_file, mode='r') as f:
                    data = json.load(f)
                    for item in data:
                        return item if item.get('ip') == ip else None
        except Exception as err:
            # self.rich.errPrint(f"加载缓存文件错误：{str(err)}")
            return None

    def clear_cache_data(self) -> None:
        """
        清理设备的缓存信息
        """
        with open(self.cache_data_file, mode='w') as f:
            f.truncate(0)
        self.rich.print("设备缓存数据已清除!")

    def connectHandler(self, host: dict) -> ConnectHandler:
        """定义一个netmiko对象"""
        try:
            connect = ''
            # 判断使用telnet协议
            if host['port'] == 23:
                # fast_cli=False，可修复telnet login authentication报错.
                host['device_type'] = f"{host['device_type']}_telnet"
                host = self.normalize_netmiko(host)
                connect = ConnectHandler(**host, fast_cli=False)
            # 判断使用ssh协议
            else:
                host = self.normalize_netmiko(host)
                if 'huawei' in host['device_type']:
                    connect = ConnectHandler(**host, conn_timeout=15)
                elif 'fortinet' in host['device_type']:
                    # 调用重写的MyFortinetSSH类
                    connect = FortinetSSH(**host)
                elif 'juniper' in host['device_type']:
                    # 优化netscreen设备(优化分屏命令)
                    connect = JuniperSSH(**host)
                elif 'maipu' in host['device_type']:      # 支持迈普厂商
                    connect = MaipuSSH(**host)
                elif 'fiberhome' in host['device_type']:  # 支持烽火厂商
                    connect = FiberHomeSSH(**host)
                elif 'hillstone' in host['device_type']:  # 支持山石厂商
                    connect = HillStoneSSH(**host)
                else:
                    connect = ConnectHandler(**host)
            return connect

        # 异常捕获
        except NetMikoAuthenticationException:
            output = "Failed.....{:<15}  用户名或密码认证失败!".format(host['ip'])
            raise ValueError(output)
        except NetmikoTimeoutException:
            output = "Failed.....{:<15}  连通性问题!".format(host['ip'])
            raise ValueError(output)
        except Exception as e:
            output = "Failed.....{:<15} 连接失败!: {}".format(host['ip'], e)
            raise ValueError(output)

    @async_task  # 等价于 run_t = async_task(run_t)
    def run_t(self, host: list) -> None:
        """
        主要获取设备名称提示符
        """
        self.rich.print('{} Start connecting device...{:<15}'.format(self.time_now(), host['ip']))
        try:
            conn = self.connectHandler(host)
            output = "获取设备的提示符: {}".format(conn.find_prompt())
            self.rich.print(output)
            self.success.append(host['ip'])  # 追加到成功的列表
            self.write_to_file(**{
                'code': 1,
                'result': output,
                'path': os.path.join(self.log, f"connect_test_{self.logtime}.log")
            })

            # 退出netmiko session
            conn.disconnect()

        except Exception as e:
            self.rich.errPrint(e)
            self.fail.append(host['ip'])
            self.write_to_file(**{
                'code': 0,
                'result': str(e),
                'path': os.path.join(self.log, f"connect_test_{self.logtime}.log")
            })

    @async_task
    def run_cmd(self, host: dict, action=1) -> None:
        """获取设备配置"""
        self.rich.print('设备...{:.<15}...开始执行'.format(host['ip']))
        try:
            # show命令条目
            cmds = host['cmd_list']
            # 特权功能标识位
            enable = True if host['secret'] else False
            # netmiko对象
            conn = self.connectHandler(host)
            # 设备名称
            hostname = self.format_hostname(conn.find_prompt(), host['device_type'])
            # 存放目录名称
            dirname = host['ip'] + '_' + hostname  # 192.168.1.1_Router-01
            # 创建目录
            dirpath = os.path.join(self.log, self.show_dir, self.logtime, dirname)
            self.create_dir(dirpath)

            if cmds:
                for cmd in cmds:
                    if enable:
                        # 进入特权模式
                        if 'huawei' or 'hp_comware' in host['device_type']:
                            # 默认源码只有cisco支持enabel命令，国产的super需要改写
                            conn.enable()
                            output = conn.send_command(cmd)
                            data = {'action': action, 'code': 1, 'result': output,
                                    'path': os.path.join(dirpath, self.format_cmd(cmd) + '.conf')}
                            self.write_to_file(**data)
                        else:
                            conn.enable()
                            output = conn.send_command(cmd)
                            data = {'action': action, 'code': 1, 'result': output,
                                    'path': os.path.join(dirpath, self.format_cmd(cmd) + '.conf')}
                            self.write_to_file(**data)
                    else:
                        output = conn.send_command(cmd)
                        data = {'action': action, 'code': 1, 'result': output,
                                'path': os.path.join(dirpath, self.format_cmd(cmd) + '.conf')}
                        self.write_to_file(**data)
            else:
                # 适用于ftp/sftp/scp备份
                if host['device_type'] == 'fortinet':
                    # 飞塔防火墙FTP备份
                    cmd = "execute backup config ftp {}_{}.conf {} {} {}".format(
                        hostname,
                        self.logtime,
                        self.FtpServer,
                        self.FtpUser, self.FtpPassword
                    )
                    conn.send_command(cmd, expect_string="to ftp server OK")
                elif host['device_type'] == 'cisco_wlc':
                    # 按需补充
                    pass
                else:
                    pass

            self.success.append(host['ip'])

        except Exception as e:
            output = f"run Failed...{host['ip']} : {e}"
            self.rich.errPrint(output)
            self.fail.append(host['ip'])
            self.write_to_file({'action': action, 'code': 0, 'result': str(e)})

        finally:
            # 退出netmiko session
            conn.disconnect()

    # @async_task
    def run_config_cmd(self, host: dict, action=1) -> None:
        """
        执行配置命令并保存当前配置
        """
        # self.printPretty('设备...{:.<15}...开始执行配置'.format(host['ip']))
        self.rich.print('设备...{:.<15}...开始执行配置'.format(host['ip']))

        # 特权功能标识位
        enable = True if host['secret'] else False
        #
        conn = self.connectHandler(host)

        if conn:
            # 获取设备名称并格式化
            hostname = self.format_hostname(conn.find_prompt(), host['device_type'])
            # 目录格式：LOG/2022-01-01_00:00:00
            dirpath = os.path.join(self.log, self.logtime)
            # 创建多级目录
            self.create_dir(dirpath)

            try:
                cmds_str = self.render_jinjia2_tpl(
                    self.normalize_vendor(host['device_type']),
                    host['template_file']
                )
                cmds = [cmd.strip() for cmd in cmds_str.splitlines()]

                output = ''
                for cmd in cmds:
                    if enable:
                        # 进入特权模式
                        if 'cisco' or 'ruijie' in host['device_type']:
                            # 默认源码只有cisco支持enabel命令，国产的super需要改写
                            conn.enable()
                            output += conn.send_config_set(cmd, exit_config_mode=False)
                        else:
                            # 占位，其他厂商
                            pass
                    else:
                        output += conn.send_config_set(cmd, exit_config_mode=False)

                # 退出配置模式
                output += conn.exit_config_mode()
                # 保存设备当前配置
                output += conn.save_config()
                # self.printPretty(output)
                self.rich.print(output)

                # 写入文件
                data = {
                    'action': action,
                    'code': 1,
                    'result': output,
                    'path': os.path.join(dirpath, f'{hostname}_Config_Record.txt')
                }
                self.write_to_file(**data)
                # 追加到成功列表
                self.success.append(host['ip'])

            except Exception as e:
                # self.printPretty(f"run_config_failed...{host['ip']} : {e}")
                self.rich.errPrint(f"run_config_failed...{host['ip']} : {e}")
                # 写入文件
                self.write_to_file(**{'action': action, 'code': 0, 'result': host['ip'] + "__" + str(e)})
                # 追加到失败列表
                self.fail.append(host['ip'])

            finally:
                # 退出netmiko session
                conn.disconnect()

    def run_ping(self) -> None:
        """ping 目标IP地址"""
        pass

    def execute_connect(self) -> None:
        """
        连接测试
        """
        # 设备信息(excel)
        host = self.get_devices_info_async()  # 列表嵌套字典
        self.run_t(host)

    def execute_getConfig(self) -> None:
        """
        下发设备配置
        """
        self.create_dir(self.show_dir)
        host = self.get_devices_info_async()  # 列表嵌套字典
        self.run_cmd(host)

    def execute_sendConfig(self) -> None:
        """
        下发设备配置
        """
        pass

    def execute_saveConfig(self) -> None:
        """
        保存设备配置
        """
        pass

    def execute_ping(self) -> None:
        """
        ping 测试
        """
        pass


if __name__ == '__main__':
    # 引用assert断言，判断一个表达式为false，抛出异常，不会往下执行
    # start_time = datetime.now()
    # result = Ping.run_ping(host='223.5.5.5')
    # end_time = datetime.now()
    # print("耗时：{:0.2f}".format((end_time - start_time).total_seconds()))

    net = NetAutoTool()
    net.execute_getConfig()

